"""Service d'export de donnees."""

from __future__ import annotations

import csv
import io
import json
import logging
import zipfile
from importlib import import_module as imp_module
from typing import Any

import openpyxl
from django.apps import apps
from django.core.files.base import ContentFile
from django.db import DatabaseError, OperationalError
from django.utils import timezone
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import ExportJob
from ..registry import MODULE_REGISTRY

logger = logging.getLogger("core.transfer")


class ExporterService:
    """Service pour l'export de donnees."""

    @classmethod
    def get_model_and_serializer(cls, module: str) -> tuple[Any, Any, dict[str, Any]]:
        """Recupere le modele et serializer pour un module."""
        config = MODULE_REGISTRY.get(module)
        if not config:
            msg = f"Module '{module}' non supporte. Modules valides: {list(MODULE_REGISTRY.keys())}"
            raise ValueError(msg)

        # Get model via Django apps registry
        try:
            model_class = apps.get_model(config["app_label"], config["model_name"])
        except LookupError as e:
            msg = f"Modele non trouve: {config['app_label']}.{config['model_name']}"
            raise ValueError(msg) from e

        # Get serializer dynamically
        try:
            serializer_module = imp_module(config["serializer_module"])
            serializer_class = getattr(serializer_module, config["serializer_name"])
        except (ImportError, AttributeError) as e:
            msg = f"Serializer non trouve: {config['serializer_module']}.{config['serializer_name']}"
            raise ValueError(msg) from e

        return model_class, serializer_class, config

    @classmethod
    def _get_filtered_queryset(
        cls,
        module: str,
        filters: dict[str, Any] | None = None,
    ) -> tuple[Any, Any]:
        """Construit le queryset optimise avec filtres pour un module."""
        model_class, serializer_class, config = cls.get_model_and_serializer(module)
        queryset = model_class.objects.all()
        if config.get("select_related"):
            queryset = queryset.select_related(*config["select_related"])
        if config.get("prefetch_related"):
            queryset = queryset.prefetch_related(*config["prefetch_related"])
        if filters:
            queryset = cls._apply_filters(queryset, filters, module)
        return queryset, serializer_class

    @classmethod
    def export_to_json(
        cls,
        module: str,
        filters: dict[str, Any] | None = None,
    ) -> tuple[str, int]:
        """Exporte les donnees au format JSON."""
        queryset, serializer_class = cls._get_filtered_queryset(module, filters)

        count = queryset.count()
        serializer = serializer_class(queryset, many=True)
        data = {
            "module": module,
            "exported_at": timezone.now().isoformat(),
            "count": count,
            "data": serializer.data,
        }

        return json.dumps(data, indent=2, ensure_ascii=False, default=str), count

    @classmethod
    def export_to_csv(
        cls,
        module: str,
        filters: dict[str, Any] | None = None,
    ) -> tuple[str, int]:
        """Exporte les donnees au format CSV."""
        queryset, serializer_class = cls._get_filtered_queryset(module, filters)

        serializer = serializer_class(queryset, many=True)
        data = serializer.data

        if not data:
            return "", 0

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()

        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                if isinstance(value, (dict, list)):
                    cleaned_row[key] = json.dumps(value, ensure_ascii=False)
                else:
                    cleaned_row[key] = value
            writer.writerow(cleaned_row)

        return output.getvalue(), len(data)

    @classmethod
    def export_to_xlsx(
        cls,
        module: str,
        filters: dict[str, Any] | None = None,
    ) -> tuple[bytes, int]:
        """Exporte les donnees au format Excel."""
        queryset, serializer_class = cls._get_filtered_queryset(module, filters)

        serializer = serializer_class(queryset, many=True)
        data = serializer.data

        workbook = openpyxl.Workbook()
        active_sheet = workbook.active
        if active_sheet is None:
            sheet = workbook.create_sheet(module.capitalize())
        else:
            sheet = active_sheet
            sheet.title = module.capitalize()

        if not data:
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue(), 0

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="673C5C", end_color="673C5C", fill_type="solid")

        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_num, row_data in enumerate(data, start=2):
            for col, header in enumerate(headers, start=1):
                value = row_data.get(header)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                sheet.cell(row=row_num, column=col, value=value)

        # Auto-adjust column widths
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            column_cells_list = list(column_cells)
            if not column_cells_list:
                continue
            col_letter = get_column_letter(col_idx)
            for cell in column_cells_list:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except (ValueError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[col_letter].width = adjusted_width

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue(), len(data)

    @classmethod
    def _apply_filters(cls, queryset: Any, filters: dict[str, Any], module: str) -> Any:
        """Applique les filtres au queryset."""
        filter_mapping = {
            "category": "category__slug",
            "category_id": "category_id",
            "status": "status__name",
            "status_id": "status_id",
            "is_published": "is_published",
            "is_featured": "is_featured",
            "type": "type__name",
            "type_id": "type_id",
        }

        module_filters = {
            "articles": ["category", "is_published", "is_featured"],
            "projects": ["category", "status"],
            "stacks": ["category"],
            "experiences": ["type"],
            "contacts": [],
        }

        allowed_filters = module_filters.get(module, [])

        for key, value in filters.items():
            if key in ("format", "page", "page_size", "limit", "offset"):
                continue

            if key in allowed_filters and key in filter_mapping and value:
                filter_key = filter_mapping[key]
                filter_value = value
                if key in ("is_published", "is_featured") and isinstance(value, str):
                    filter_value = value.lower() in ("true", "1", "yes", "oui")
                queryset = queryset.filter(**{filter_key: filter_value})

        return queryset

    @classmethod
    def create_export_job(
        cls,
        user: Any,
        module: str,
        export_format: str,
        filters: dict[str, Any] | None = None,
    ) -> ExportJob:
        """Cree et execute un job d'export."""
        if module not in MODULE_REGISTRY:
            msg = f"Module '{module}' non supporte. Modules valides: {list(MODULE_REGISTRY.keys())}"
            raise ValueError(msg)

        valid_formats = [choice[0] for choice in ExportJob.Format.choices]
        if export_format not in valid_formats:
            msg = f"Format '{export_format}' non supporte. Formats valides: {valid_formats}"
            raise ValueError(msg)

        job = ExportJob.objects.create(
            user=user,
            module=module,
            format=export_format,
            filters=filters or {},
            status=ExportJob.Status.PENDING,
        )

        return cls.run_export(job)

    @classmethod
    def run_export(cls, job: ExportJob) -> ExportJob:
        """Execute l'export pour un job existant."""
        job.status = ExportJob.Status.PROCESSING
        job.save(update_fields=["status"])

        try:
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            export_format = job.format
            module = job.module
            filters = job.filters
            count = 0

            if export_format == ExportJob.Format.JSON:
                text_content, count = cls.export_to_json(module, filters)
                filename = f"{module}_{timestamp}.json"
                job.file.save(filename, ContentFile(text_content.encode("utf-8")))

            elif export_format == ExportJob.Format.CSV:
                text_content, count = cls.export_to_csv(module, filters)
                filename = f"{module}_{timestamp}.csv"
                job.file.save(filename, ContentFile(text_content.encode("utf-8")))

            elif export_format == ExportJob.Format.XLSX:
                bytes_content, count = cls.export_to_xlsx(module, filters)
                filename = f"{module}_{timestamp}.xlsx"
                job.file.save(filename, ContentFile(bytes_content))

            job.status = ExportJob.Status.COMPLETED
            job.records_count = count
            job.completed_at = timezone.now()
            job.save()

        except (DatabaseError, OperationalError, ValueError, ImportError):
            logger.exception("Erreur lors de l'export du module %s", job.module)
            job.status = ExportJob.Status.FAILED
            job.error_message = "Erreur lors de l'export"
            job.save()

        return job

    @classmethod
    def export_multiple_to_zip(
        cls,
        modules: list[str],
        export_format: str,
        filters: dict[str, Any] | None = None,
    ) -> tuple[bytes, dict[str, int]]:
        """Exporte plusieurs modules dans un fichier ZIP.

        Returns:
            Tuple (zip_content, module_counts)
        """
        module_counts: dict[str, int] = {}
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        files_added = 0

        # Normaliser le format (case-insensitive)
        export_format = export_format.lower().strip()

        logger.info(
            "=== Debut export ZIP: modules=%s, format=%s ===",
            modules,
            export_format,
        )

        # Creer le buffer et le ZIP
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            for module in modules:
                if module not in MODULE_REGISTRY:
                    logger.warning("Module '%s' non supporte, ignore", module)
                    continue

                try:
                    file_content: bytes
                    filename: str
                    count: int = 0

                    logger.info("Export module %s au format %s...", module, export_format)

                    if export_format == "json":
                        content, count = cls.export_to_json(module, filters)
                        if not content:
                            logger.warning("Export JSON vide pour %s", module)
                            content = '{"module": "' + module + '", "data": []}'
                        filename = f"{module}_{timestamp}.json"
                        file_content = content.encode("utf-8")

                    elif export_format == "csv":
                        content, count = cls.export_to_csv(module, filters)
                        if not content:
                            logger.warning("Export CSV vide pour %s", module)
                            content = ""
                        filename = f"{module}_{timestamp}.csv"
                        file_content = content.encode("utf-8")

                    elif export_format == "xlsx":
                        xlsx_content, count = cls.export_to_xlsx(module, filters)
                        if not xlsx_content:
                            logger.warning("Export XLSX vide pour %s", module)
                            continue
                        filename = f"{module}_{timestamp}.xlsx"
                        file_content = xlsx_content

                    else:
                        logger.error(
                            "Format '%s' non supporte pour le module %s",
                            export_format,
                            module,
                        )
                        continue

                    # Verifier que le contenu n'est pas vide
                    if len(file_content) == 0:
                        logger.warning("Contenu vide pour %s, skip", module)
                        continue

                    # Ecrire le fichier dans le ZIP
                    logger.info(
                        "Ajout de %s au ZIP (%d bytes, %d records)",
                        filename,
                        len(file_content),
                        count,
                    )
                    zip_file.writestr(filename, file_content)
                    files_added += 1
                    module_counts[module] = count

                except (ValueError, ImportError, LookupError):
                    logger.exception("Erreur export module %s", module)
                    module_counts[module] = -1
                except (DatabaseError, OperationalError):
                    logger.exception("Erreur DB export module %s", module)
                    module_counts[module] = -1

        # Verifier que des fichiers ont ete ajoutes
        if files_added == 0:
            logger.error("ATTENTION: Aucun fichier ajoute au ZIP!")

        # Recuperer le contenu avec getvalue()
        zip_bytes = zip_buffer.getvalue()

        logger.info(
            "=== ZIP finalise: %d fichiers, %d bytes total ===",
            files_added,
            len(zip_bytes),
        )

        return zip_bytes, module_counts
