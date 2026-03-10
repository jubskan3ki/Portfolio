"""Excel (XLSX) import/export strategy."""

import io
import json
from typing import Any

import openpyxl
from django.core.files.uploadedfile import UploadedFile
from openpyxl.utils import get_column_letter

from .base import ExportStrategy, ImportStrategy


class XlsxImportStrategy(ImportStrategy):
    """Strategy for importing Excel files."""

    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def file_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def parse(self, file: UploadedFile) -> list[dict[str, Any]]:
        """Parse Excel file."""
        workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        records: list[dict[str, Any]] = []

        if sheet is None:
            workbook.close()
            return records

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            workbook.close()
            return records

        # Build headers from first row
        headers = self._build_headers(rows[0])

        # Parse data rows
        for row in rows[1:]:
            record = self._parse_row(row, headers)
            if record:  # Only add non-empty records
                records.append(record)

        workbook.close()
        return records

    def _build_headers(self, header_row: tuple) -> list[str]:
        """Build headers from first row."""
        headers = []
        for i, h in enumerate(header_row):
            if h is not None:
                headers.append(str(h).strip())
            else:
                headers.append(f"col_{i}")
        return headers

    def _parse_row(self, row: tuple, headers: list[str]) -> dict[str, Any] | None:
        """Parse a data row into a dictionary."""
        record = {}
        for i, value in enumerate(row):
            if i >= len(headers):
                continue

            header = headers[i]

            # Try to parse JSON strings
            if isinstance(value, str) and value.startswith(("[", "{")):
                try:
                    record[header] = json.loads(value)
                except json.JSONDecodeError:
                    record[header] = value
            else:
                record[header] = value

        # Return None for empty records
        if not any(v is not None and v != "" for v in record.values()):
            return None

        return record


class XlsxExportStrategy(ExportStrategy):
    """Strategy for exporting Excel files."""

    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    def serialize(self, data: list[dict[str, Any]], module: str) -> bytes:
        """Serialize data to Excel."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if sheet is None:
            sheet = workbook.create_sheet()

        sheet.title = module.capitalize()

        if not data:
            output = io.BytesIO()
            workbook.save(output)
            return output.getvalue()

        # Collect all unique keys
        all_keys: set[str] = set()
        for record in data:
            all_keys.update(record.keys())

        headers = sorted(all_keys)

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data
        for row_idx, record in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = record.get(header, "")

                # Convert complex objects to JSON strings
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)

                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))

            for row in range(2, min(len(data) + 2, 100)):  # Limit to first 100 rows for performance
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))

            sheet.column_dimensions[column_letter].width = max_length + 2

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()
